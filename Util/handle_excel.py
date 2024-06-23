#coding=utf-8
import openpyxl
from openpyxl import workbook
import sys
from collections.abc import Iterable
import os
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../"))
sys.path.append(base_path)
casePath = os.path.abspath(os.path.join(os.path.dirname(__file__), "../Case"))


class HandExcel:
    def load_excel(self, filename):
        '''
        加载excel
        '''
        open_excel = openpyxl.load_workbook(casePath+"/{}".format(filename))
        # print(open_excel)
        return open_excel

    def get_sheet_data(self, filename, index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel(filename).sheetnames
        if index == None:
            index = 0
        data = self.load_excel(filename)[sheet_name[index]]
        return data

    def get_cell_value(self, filename, row, cols):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data(filename).cell(row=row,column=cols).value
        return data

    def get_rows(self, filename):
        '''
        获取行数
        '''
        row = self.get_sheet_data(filename).max_row
        return row - 1

    def get_rows_value(self, filename, row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data(filename)[row]:
            row_list.append(i.value)
        return row_list


    def excel_write_data(self, filename, row, cols, value):
        '''
        写入数据
        '''
        wb = self.load_excel(filename)
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(casePath + "/{}".format(filename))


    def get_columns_value(self, filename, key=None):
        '''
        获取某一列的数据
        '''
        columns_list = []
        if key==None:
            key = 'A'
        columns_list_data = self.get_sheet_data(filename)[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, filename, case_id):
        '''
        获取行号
        '''
        num = 1
        cols_data = self.get_columns_value(filename)
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num+1
        return num

    def get_excel_data(self, filename):
        '''
        获取excel里面所有的数据
        '''
        data_list = []
        for i in range(self.get_rows(filename)):
            data_list.append(self.get_rows_value(filename, i+2))
        return data_list


excel_data = HandExcel()

if __name__ == '__main__':
    handle = HandExcel()
    filename = "icash_HK_api_testcase.xlsx"
    print(handle.get_rows_number(filename,'cashorder_001'))
    print(handle.get_excel_data(filename))
    print(handle.get_columns_value(filename, 'A'))
    print(handle.get_rows_number(filename, 'imooc_001'))
    print(handle.get_rows_value(filename,3))
